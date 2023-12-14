import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url='http://news.v.daum.net/v/20211129144552297' #:뉴스 페이지에 접속합니다. 네이버, 다음 등에서 아무런 뉴스를 클릭하고 접속한 주소를 입력합니다.

headers={ #7~10: 헤더정보를 입력합니다. 헤더정보를 입력하지 않을 시 사이트에서 로봇이 접속한 것으로 판단하여 응답하지 않는 사이트가 많습니다.
  'User-Agent':'Mozilla/5.0',
  'Content-Type': 'text/html; charset=utf-8'
  }

response=requests.get(url,headers=headers)

results=re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results=list(set(results))

print(results)

try:
  wb=load_workbook(r"13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx", data_only=True)
  sheet=wb.active
except:
  wb=Workbook()
  sheet=wb.active

for result in results:
  sheet.append([result])

wb.save(r"13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx")